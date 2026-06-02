import os
import time
from collections import deque
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from threading import Thread
from loguru import logger
import query_main
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

logger.add("log/log_new.log", rotation="100 MB", retention="1 week", encoding="utf-8",
           format="<green>{time:YYYY-MM-DD HH:mm:ss.SSS}</green> | <level>{level: <8}</level> | Thread: {thread} | <cyan>{name}</cyan>:<cyan>{function}</cyan>:<cyan>{line}</cyan> - <level>{message}</level>",
           )


def get_filename_name(filepath):

    filenames = [i for i in os.listdir(filepath) if "~" not in i]
    # print(len(filenames))
    # index = 0
    if len(filenames) == 0:
        print(f'当前目录为：{filepath},未检测到输入文件。')
        input("输入任意内容结束")
        exit()

    if len(filenames) == 1:
        print('只检测到一个文件名，直接返回')
        return os.path.join(filepath, filenames[0]), filenames[0]

    print('请选择输入文件')
    for index,filename in enumerate(filenames):
        print(f"{index}: {filename}")

    while True:
        try:
            index = int(input("请输入文件前的序号"))
        except:
            print('输入错误，请重新输入，请确保输入的是数字')
            continue

        # 检测数字是否合法，合法就结束输入，否则继续输入
        if index<0 or index>=len(filenames):
            print("数字不合法，超出可取范围，请重新输入")
        else:
            break

    # if len(filenames) != 1:
    #     raise FileExistsError(f"文件数量异常。只允许为1。目前为：{len(filenames)}")
    return os.path.join(filepath, filenames[index]), filenames[index]


def read_txt_file(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            # 读取文件的每一行并存储在列表中
            lines = file.readlines()
            # 去除每行末尾的换行符
            lines = [line.strip() for line in lines]
            return lines
    except FileNotFoundError:
        print(f"文件 {file_path} 未找到。")
        return []


def read_txt_with_extract(file_path):
    """
    读取包含"手机号----姓名----身份证号"格式的txt文件
    返回二维列表[[姓名, 身份证号], ...]
    """
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            result = []
            for line in file:
                line = line.strip()
                if not line:
                    continue
                parts = line.split('----')
                if len(parts) >= 3:  # 兼容可能多余分隔符的情况
                    result.append([parts[1].strip(), parts[2].strip()])
                    # logger.info(f"成功解析：{parts[1]} - {parts}")
                else:
                    # pass
                    logger.warning(f"格式错误行：{line}")
            return result
    except FileNotFoundError:
        print(f"文件 {file_path} 未找到。")
        return []


def run():
    input_path = "input"
    output_path = 'output'
    # max_workers = 1
    if not os.path.exists(output_path):
        os.makedirs(output_path)
    if not os.path.exists(input_path):
        os.makedirs(input_path)

    try:
        mode = input("请选择手机号或身份证查询方式：1. 手机号 2. 身份证")
        if int(mode) not in [1,2]:
            input("输入错误")
            exit()
        mode = int(mode)
    except:
        input("输入错误")
        exit()

    # mode = 2

    try:
        proxies_mode = input("请选择代理模式：1. 提取式代理 2. 隧道代理")
        if int(proxies_mode) not in [1,2]:
            input("输入错误")
            exit()
    except:
        input("输入错误")
        exit()

    proxies_url = input("请输入代理链接：")

    max_workers = int(input("请输入线程数："))
    if mode == 1:
        area = input("请输入区号(不带+号)：")
    else:
        pass
    # 读取文件
    input_file, filename = get_filename_name(input_path)
    output_name = f'{output_path}/{filename.split(".")[0]}-{datetime.strftime(datetime.now(), "%Y-%m-%d_%H-%M-%S")}.xlsx'

    result_dict = {}
    phone_list = read_txt_file(input_file)

    init_queue = deque()

    if mode == 1:
        for i in phone_list:
            if i.strip():
                init_queue.appendleft(("+"+area, i))
    else:
        id_list = read_txt_with_extract(input_file)
        # id_list 的结构是 [[于子川, 230102199011282835], ...]
        for item in id_list:
            init_queue.appendleft((item[0], item[1]))  # 姓名和身份证号

    executor = ThreadPoolExecutor(max_workers=max_workers)
    for i in range(max_workers*5):
        executor.submit(query_main.run, init_queue, result_dict, proxies_mode, proxies_url,user_card=False if mode == 1 else True)

    index = 0
    while True:
        if result_dict:
            temp_dict = result_dict.copy()
            # 检查文件是否存在，如果不存在则创建一个新的 Excel 文件
            if not os.path.exists(output_name):
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                if mode == 1:
                    ws.append(["area", 'phone', 'user_name', 'year', 'status', "is_real_name"])
                else:
                    ws.append(["name", 'idcard', 'user_name', 'year','status',"is_real_name"])
            else:
                wb = load_workbook(output_name)
                ws = wb.active

            for i in temp_dict:
                temp_output_data = result_dict.pop(i)
                area = temp_output_data.get("area", '无')
                phone = temp_output_data.get("phone", '无')
                user_name = temp_output_data.get("user_name", '无')
                year = temp_output_data.get("year", '无')
                status = temp_output_data.get("status", '无')
                is_real_name = temp_output_data.get("is_real_name", '无')

                ws.append([area, phone, user_name, year, status, is_real_name])
                index += 1
                logger.success(f"已输出：{index}/{len(phone_list)}")
            while True:
                try:
                    wb.save(output_name)
                    wb.close()
                    logger.success(f"文件保存成功：{output_name}")
                    break
                except Exception as e:
                    logger.warning(f"文件被占用，正在重试:{e}")
                    time.sleep(1)
        else:
            time.sleep(1)


if __name__ == '__main__':
    # print(read_txt_with_extract("input/lz.txt"))
    run()

